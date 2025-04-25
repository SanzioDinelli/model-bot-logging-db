from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, Alignment, PatternFill
from openpyxl.chart import BarChart, Reference
from openpyxl.drawing.image import Image
from openpyxl.formatting.rule import CellIsRule
from openpyxl.worksheet.protection import SheetProtection
from copy import copy
import logging

def main_excel():
    # === 1. Criar um arquivo Excel do zero ===
    wb = Workbook()
    ws = wb.active
    ws.title = "Vendas"

    # Cabeçalhos
    ws.append(["Produto", "Quantidade", "Preço Unitário", "Total"])

    # Dados
    dados = [
        ["Caneta", 10, 2.5],
        ["Caderno", 5, 10.0],
        ["Borracha", 10, 0.9],
    ]

    for linha in dados:
        total = linha[1] * linha[2]
        ws.append(linha + [total])

    # === 2. Modificar células: valores, fórmulas, estilos, alinhamento ===
    ws["A1"].font = Font(bold=True, color="FFFFFF")
    ws["A1"].fill = PatternFill(start_color="4F81BD", fill_type="solid")
    ws["A1"].alignment = Alignment(horizontal="center")

    # Aplicar o mesmo estilo para outros cabeçalhos
    for cell in ws[1]:
        cell.font = Font(bold=True)
        cell.fill = PatternFill(start_color="4F81BD", fill_type="solid")
        cell.alignment = Alignment(horizontal="center")

    # === 3. Trabalhar com múltiplas abas ===
    ws2 = wb.create_sheet(title="Resumo")
    ws2["A1"] = "Planilha criada como resumo"

    # === 4. Inserir gráficos ===
    chart = BarChart()
    chart.title = "Totais por Produto"
    chart.x_axis.title = "Produto"
    chart.y_axis.title = "Total"

    data = Reference(ws, min_col=4, min_row=1, max_row=ws.max_row)
    cats = Reference(ws, min_col=1, min_row=2, max_row=ws.max_row)
    chart.add_data(data, titles_from_data=True)
    chart.set_categories(cats)
    ws.add_chart(chart, "F2")

    # === 5. Inserir imagens ===
    img = Image("logo.png")  # Certifique-se que este arquivo exista na mesma pasta
    img.width = 100
    img.height = 100
    ws.add_image(img, "H1")

    # === 6. Aplicar formatação condicional ===
    # Destacar em vermelho totais menores que 20
    red_fill = PatternFill(start_color="FFC7CE", fill_type="solid")
    ws.conditional_formatting.add(
        "D2:D4",
        CellIsRule(operator='lessThan', formula=[20], fill=red_fill)
    )

    # === 7. Proteger planilhas e células ===
    # Proteger toda a planilha
    ws.protection = SheetProtection(sheet=True, password="12345")

    # Desproteger células específicas (por exemplo: cabeçalhos)
    for cell in ws["1:1"]:
        protection = copy(cell.protection)
        protection.locked = False
        cell.protection = protection

    # === 8. Salvar o arquivo ===
    wb.save("planilha_completa.xlsx")
    wb.close()

    # === 9. Abrir e ler planilhas existentes ===
    wb_lido = load_workbook("planilha_completa.xlsx")
    ws_lido = wb_lido["Vendas"]

    logging.debug(f'Conteúdo da célula A1:{ws_lido["A1"].value}')
    wb_lido.close()
